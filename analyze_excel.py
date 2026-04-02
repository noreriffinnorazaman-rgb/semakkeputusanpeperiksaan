import pandas as pd
import openpyxl

excel_file = "REA0084 SMKTLS SPM 2026.xlsm (2).xlsx"

# Use openpyxl to read the actual Excel file
wb = openpyxl.load_workbook(excel_file, data_only=True)

print("=== ANALYZING EXCEL STRUCTURE ===\n")

# Check I1 sheet structure
print("I1 Sheet Analysis:")
ws = wb['I1']
for row in range(1, 31):
    values = []
    for col in range(1, 6):
        cell = ws.cell(row, col)
        val = str(cell.value)[:40] if cell.value else ''
        values.append(val)
    if any(values):
        print(f"Row {row}: {values}")

print("\n\nB1 Sheet Analysis:")
ws = wb['B1']
for row in range(1, 11):
    values = []
    for col in range(1, 6):
        cell = ws.cell(row, col)
        val = str(cell.value)[:40] if cell.value else ''
        values.append(val)
    if any(values):
        print(f"Row {row}: {values}")

print("\n\nSENARAI NAMA A Sheet Analysis:")
ws = wb['SENARAI NAMA A']
for row in range(1, 11):
    values = []
    for col in range(1, 12):
        cell = ws.cell(row, col)
        val = str(cell.value)[:30] if cell.value else ''
        values.append(val)
    if any(values):
        print(f"Row {row}: {values}")

wb.close()
