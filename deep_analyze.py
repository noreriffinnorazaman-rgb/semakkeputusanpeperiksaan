import openpyxl
import json

excel_file = "REA0084 SMKTLS SPM 2026.xlsm (2).xlsx"
wb = openpyxl.load_workbook(excel_file, data_only=True)

# ============================================================
# DEEP ANALYSIS OF EVERY SHEET
# ============================================================

# 1. Analyze I1 sheet completely (student template)
print("="*80)
print("I1 SHEET - COMPLETE DUMP (Student Template)")
print("="*80)
ws = wb['I1']
for row in range(1, ws.max_row + 1):
    vals = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# 2. Check which I sheets have student names
print("\n" + "="*80)
print("ALL I SHEETS - STUDENT NAMES CHECK")
print("="*80)
for i in range(1, 51):
    sn = f'I{i}'
    if sn in wb.sheetnames:
        ws = wb[sn]
        # Check multiple possible locations for student name
        name_candidates = []
        for r in range(1, 10):
            for c in range(1, 10):
                v = ws.cell(r, c).value
                if v and str(v).strip() and str(v) not in ['HEADCOUNT CEMERLANG', "HEADCOUNT CEMERLANG 'A'", '2026', 'NAMA MURID :', 'KELAS :', 'NO. K/P :', 'JANTINA :', 'BIL.', 'MATA PELAJARAN', 'TOV', 'U1', 'PPT', 'SPMC', 'ETR', 'M', 'MATA', 'GRED', 'nan']:
                    name_candidates.append(f"R{r}C{c}={v}")
        print(f"{sn}: {name_candidates[:10]}")

# 3. Analyze HEADCOUNT sheet completely
print("\n" + "="*80)
print("HEADCOUNT SHEET - COMPLETE")
print("="*80)
ws = wb['HEADCOUNT']
for row in range(1, min(ws.max_row + 1, 60)):
    vals = []
    for col in range(1, min(ws.max_column + 1, 20)):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# 4. Analyze HC KESELURUHAN
print("\n" + "="*80)
print("HC KESELURUHAN SHEET")
print("="*80)
ws = wb['HC KESELURUHAN']
for row in range(1, min(ws.max_row + 1, 50)):
    vals = []
    for col in range(1, min(ws.max_column + 1, 20)):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# 5. Analyze ANALISIS - KPI
print("\n" + "="*80)
print("ANALISIS - KPI SHEET")
print("="*80)
ws = wb['ANALISIS - KPI']
for row in range(1, min(ws.max_row + 1, 50)):
    vals = []
    for col in range(1, min(ws.max_column + 1, 20)):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# 6. Analyze ANALISIS - HC
print("\n" + "="*80)
print("ANALISIS - HC SHEET")
print("="*80)
ws = wb['ANALISIS - HC']
for row in range(1, min(ws.max_row + 1, 50)):
    vals = []
    for col in range(1, min(ws.max_column + 1, 20)):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# 7. Analyze B1 sheet completely (subject headcount template)
print("\n" + "="*80)
print("B1 SHEET - COMPLETE DUMP (Subject Headcount Template)")
print("="*80)
ws = wb['B1']
for row in range(1, ws.max_row + 1):
    vals = []
    for col in range(1, min(ws.max_column + 1, 25)):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# 8. Analyze SENARAI NAMA A
print("\n" + "="*80)
print("SENARAI NAMA A SHEET - COMPLETE")
print("="*80)
ws = wb['SENARAI NAMA A']
for row in range(1, ws.max_row + 1):
    vals = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# 9. Analyze A. CEMERLANG A
print("\n" + "="*80)
print("A. CEMERLANG A SHEET")
print("="*80)
ws = wb['A. CEMERLANG A']
for row in range(1, min(ws.max_row + 1, 30)):
    vals = []
    for col in range(1, min(ws.max_column + 1, 20)):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# 10. Analyze A. LAYAK SIJIL
print("\n" + "="*80)
print("A. LAYAK SIJIL SHEET")
print("="*80)
ws = wb['A. LAYAK SIJIL']
for row in range(1, min(ws.max_row + 1, 30)):
    vals = []
    for col in range(1, min(ws.max_column + 1, 20)):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# 11. Analyze MAKLUMAT ASAS completely
print("\n" + "="*80)
print("MAKLUMAT ASAS SHEET - COMPLETE")
print("="*80)
ws = wb['MAKLUMAT ASAS']
for row in range(1, ws.max_row + 1):
    vals = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# 12. Check HC-TOV, HC-U1, HC-PPT, HC-SPMC, HC-ETR
for hc_sheet in ['HC-TOV', 'HC-U1', 'HC-PPT', 'HC-SPMC', 'HC-ETR']:
    if hc_sheet in wb.sheetnames:
        print(f"\n{'='*80}")
        print(f"{hc_sheet} SHEET")
        print("="*80)
        ws = wb[hc_sheet]
        for row in range(1, min(ws.max_row + 1, 30)):
            vals = []
            for col in range(1, min(ws.max_column + 1, 15)):
                v = ws.cell(row, col).value
                if v is not None:
                    vals.append(f"C{col}={v}")
            if vals:
                print(f"Row {row}: {vals}")

wb.close()
print("\n\nDONE - Deep analysis complete!")
