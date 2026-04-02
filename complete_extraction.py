import openpyxl
import json

def extract_complete_data(excel_file):
    """Extract ALL data from Excel file"""
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    all_data = {
        'school_info': {},
        'students': [],
        'subjects': {
            'bahasa': [],
            'sains': [],
            'kemanusiaan': [],
            'vokasional': []
        },
        'headcount': {},
        'analysis': {}
    }
    
    # Extract school information from MAKLUMAT ASAS
    print("Extracting school information...")
    ws = wb['MAKLUMAT ASAS']
    all_data['school_info'] = {
        'school_name': ws.cell(4, 4).value or 'SMK TUANKU LAILATUL SHAHREEN',
        'school_code': ws.cell(4, 10).value or 'REA0084',
        'principal': ws.cell(6, 4).value or '',
        'grade': ws.cell(6, 10).value or 'B',
        'exam_year': ws.cell(10, 4).value or 2026,
        'form': ws.cell(12, 4).value or 'LIMA'
    }
    
    # Extract all students from I sheets (I1 to I50)
    print("\nExtracting student data from I sheets...")
    for i in range(1, 51):
        sheet_name = f'I{i}'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Get student name from row 7
            student_name = ws.cell(7, 2).value
            
            if student_name and student_name != 'NAMA MURID :':
                student = {
                    'id': i,
                    'sheet': sheet_name,
                    'name': student_name,
                    'class': ws.cell(8, 2).value or '',
                    'ic': ws.cell(9, 2).value or '',
                    'subjects': []
                }
                
                # Extract subjects (rows 12-23)
                subject_list = [
                    'BAHASA MELAYU', 'BAHASA INGGERIS', 'SEJARAH', 'MATEMATIK',
                    'SAINS', 'PENDIDIKAN ISLAM', 'PRINSIP AKAUAN', 'MATEMATIK TAMBAHAN'
                ]
                
                for row in range(12, 24):
                    subject_name = ws.cell(row, 2).value
                    if subject_name and subject_name not in ['JUMLAH GRED', 'JUMLAH MATA', 'JUMLAH M.PELAJARAN', 'GRED PURATA MURID']:
                        subject = {
                            'name': subject_name,
                            'tov_marks': ws.cell(row, 4).value or 0,
                            'tov_grade': ws.cell(row, 5).value or '',
                            'u1_marks': ws.cell(row, 7).value or 0,
                            'u1_grade': ws.cell(row, 8).value or '',
                            'ppt_marks': ws.cell(row, 10).value or 0,
                            'ppt_grade': ws.cell(row, 11).value or '',
                            'spmc_marks': ws.cell(row, 13).value or 0,
                            'spmc_grade': ws.cell(row, 14).value or '',
                            'etr_marks': ws.cell(row, 16).value or 0,
                            'etr_grade': ws.cell(row, 17).value or ''
                        }
                        student['subjects'].append(subject)
                
                # Get summary data
                student['total_grade'] = ws.cell(24, 4).value or 0
                student['total_marks'] = ws.cell(25, 4).value or 0
                student['total_subjects'] = ws.cell(26, 4).value or 8
                student['average_grade'] = ws.cell(27, 4).value or 0
                
                all_data['students'].append(student)
                print(f"  Student {i}: {student_name}")
    
    # Extract Bahasa subjects (B1-B8)
    print("\nExtracting Bahasa subjects...")
    for i in range(1, 9):
        sheet_name = f'B{i}'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            subject_name = ws.cell(8, 2).value
            if subject_name:
                subject_data = extract_subject_sheet(ws, sheet_name, subject_name)
                all_data['subjects']['bahasa'].append(subject_data)
                print(f"  {sheet_name}: {subject_name}")
    
    # Extract Sains subjects (S1-S8)
    print("\nExtracting Sains subjects...")
    for i in range(1, 9):
        sheet_name = f'S{i}'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            subject_name = ws.cell(8, 2).value
            if subject_name:
                subject_data = extract_subject_sheet(ws, sheet_name, subject_name)
                all_data['subjects']['sains'].append(subject_data)
                print(f"  {sheet_name}: {subject_name}")
    
    # Extract Kemanusiaan subjects (K1-K14)
    print("\nExtracting Kemanusiaan subjects...")
    for i in range(1, 15):
        sheet_name = f'K{i}'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            subject_name = ws.cell(8, 2).value
            if subject_name:
                subject_data = extract_subject_sheet(ws, sheet_name, subject_name)
                all_data['subjects']['kemanusiaan'].append(subject_data)
                print(f"  {sheet_name}: {subject_name}")
    
    # Extract Vokasional subjects (V1-V24)
    print("\nExtracting Vokasional subjects...")
    for i in range(1, 25):
        sheet_name = f'V{i}'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            subject_name = ws.cell(8, 2).value
            if subject_name:
                subject_data = extract_subject_sheet(ws, sheet_name, subject_name)
                all_data['subjects']['vokasional'].append(subject_data)
                print(f"  {sheet_name}: {subject_name}")
    
    wb.close()
    return all_data

def extract_subject_sheet(ws, sheet_name, subject_name):
    """Extract data from a subject sheet"""
    data = {
        'sheet': sheet_name,
        'name': subject_name,
        'school': ws.cell(6, 2).value or '',
        'statistics': []
    }
    
    # Extract grade statistics (rows 10 onwards)
    row = 11
    while row < 100:
        grade = ws.cell(row, 1).value
        if not grade or grade == 'JUMLAH':
            break
        
        stat = {
            'achievement': grade,
            'registered': ws.cell(row, 2).value or 0,
            'taken': ws.cell(row, 3).value or 0,
            'grade': ws.cell(row, 4).value or ''
        }
        data['statistics'].append(stat)
        row += 1
    
    return data

if __name__ == "__main__":
    excel_file = "REA0084 SMKTLS SPM 2026.xlsm (2).xlsx"
    
    print("="*60)
    print("COMPLETE DATA EXTRACTION - SPM 2026")
    print("="*60)
    
    data = extract_complete_data(excel_file)
    
    # Save to JSON
    with open('spm_data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    print("\n" + "="*60)
    print("EXTRACTION SUMMARY")
    print("="*60)
    print(f"School: {data['school_info']['school_name']}")
    print(f"Code: {data['school_info']['school_code']}")
    print(f"Year: {data['school_info']['exam_year']}")
    print(f"\nTotal Students: {len(data['students'])}")
    print(f"Bahasa Subjects: {len(data['subjects']['bahasa'])}")
    print(f"Sains Subjects: {len(data['subjects']['sains'])}")
    print(f"Kemanusiaan Subjects: {len(data['subjects']['kemanusiaan'])}")
    print(f"Vokasional Subjects: {len(data['subjects']['vokasional'])}")
    print(f"\nData saved to: spm_data.json")
    print("="*60)
