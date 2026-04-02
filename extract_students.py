import openpyxl
import json

# ============================================================
# EXTRACT ALL STUDENT NAMES FROM SENARAI NAMA MURID 2026
# ============================================================
print("="*80)
print("EXTRACTING STUDENT NAMES FROM SENARAI NAMA MURID2026.xlsx")
print("="*80)

wb2 = openpyxl.load_workbook("SENARAI NAMA MURID2026.xlsx", data_only=True)
print(f"Sheet names: {wb2.sheetnames}")

for sn in wb2.sheetnames:
    ws = wb2[sn]
    print(f"\n--- Sheet: {sn} (rows={ws.max_row}, cols={ws.max_column}) ---")
    for row in range(1, min(ws.max_row + 1, 120)):
        vals = []
        for col in range(1, min(ws.max_column + 1, 15)):
            v = ws.cell(row, col).value
            if v is not None:
                vals.append(f"C{col}={v}")
        if vals:
            print(f"Row {row}: {vals}")

wb2.close()
