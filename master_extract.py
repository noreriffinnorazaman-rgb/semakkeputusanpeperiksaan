import openpyxl
import json

def safe_val(v):
    """Convert value to JSON-safe type"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    return str(v).strip()

def extract_all():
    """Extract ALL data from both Excel files into one comprehensive JSON"""
    
    data = {
        'school_info': {},
        'exam_config': {},
        'students_by_class': {},
        'subject_headcount': {},
        'headcount_overall': {},
        'hc_keseluruhan': {},
        'analysis_kpi': {},
        'analysis_hc': {},
        'cemerlang_a': {},
        'layak_sijil': {},
        'hc_by_exam': {},
        'individual_sheets': {},
        'all_subjects_list': []
    }
    
    # ================================================================
    # 1. EXTRACT ALL STUDENT NAMES FROM SENARAI NAMA MURID2026.xlsx
    # ================================================================
    print("="*70)
    print("STEP 1: Extracting ALL student names from SENARAI NAMA MURID2026.xlsx")
    print("="*70)
    
    wb2 = openpyxl.load_workbook("SENARAI NAMA MURID2026.xlsx", data_only=True)
    total_students = 0
    
    for sheet_name in wb2.sheetnames:
        ws = wb2[sheet_name]
        current_class = None
        students_in_class = []
        
        for row in range(1, ws.max_row + 1):
            c1 = ws.cell(row, 1).value
            c2 = ws.cell(row, 2).value
            c3 = ws.cell(row, 3).value
            
            # Detect class header
            if c2 and (not c1 or str(c1).strip() == '') and str(c2).strip() and str(c2).strip() not in ['Nama', 'NAMA']:
                potential_class = str(c2).strip()
                # Clean up class name (remove "KELAS : " prefix)
                if potential_class.startswith('KELAS'):
                    potential_class = potential_class.replace('KELAS :', '').replace('KELAS:', '').strip()
                if any(k in potential_class for k in ['BESTARI', 'PROGRESIF', 'KREATIF', 'DINAMIK']):
                    # Save previous class if exists
                    if current_class and students_in_class:
                        data['students_by_class'][current_class] = students_in_class
                        total_students += len(students_in_class)
                        print(f"  {current_class}: {len(students_in_class)} students")
                    current_class = potential_class
                    students_in_class = []
                    continue
            
            # Also check C3 for class name - switch class if different
            if c3 and str(c3).strip() and any(k in str(c3).strip() for k in ['BESTARI', 'PROGRESIF', 'KREATIF', 'DINAMIK']):
                c3_class = str(c3).strip()
                if c3_class != current_class:
                    # Save previous class if exists
                    if current_class and students_in_class:
                        data['students_by_class'][current_class] = students_in_class
                        total_students += len(students_in_class)
                        print(f"  {current_class}: {len(students_in_class)} students")
                    current_class = c3_class
                    students_in_class = []
            
            # Detect student row (has a number in C1 and a name in C2)
            if c1 is not None and c2 is not None:
                try:
                    bil = int(float(str(c1)))
                    name = str(c2).strip()
                    if name and name not in ['Nama', 'NAMA', 'Bil', 'BIL', '']:
                        student = {
                            'bil': bil,
                            'name': name,
                            'class': current_class or '',
                            'form': sheet_name
                        }
                        students_in_class.append(student)
                except (ValueError, TypeError):
                    pass
        
        # Save last class
        if current_class and students_in_class:
            data['students_by_class'][current_class] = students_in_class
            total_students += len(students_in_class)
            print(f"  {current_class}: {len(students_in_class)} students")
    
    wb2.close()
    print(f"\n  TOTAL STUDENTS EXTRACTED: {total_students}")
    
    # ================================================================
    # 2. EXTRACT ALL DATA FROM MAIN EXCEL FILE
    # ================================================================
    print("\n" + "="*70)
    print("STEP 2: Extracting ALL data from REA0084 SMKTLS SPM 2026.xlsm (2).xlsx")
    print("="*70)
    
    wb = openpyxl.load_workbook("REA0084 SMKTLS SPM 2026.xlsm (2).xlsx", data_only=True)
    
    # --- 2a. MAKLUMAT ASAS (School Info) ---
    print("\n  [2a] MAKLUMAT ASAS...")
    ws = wb['MAKLUMAT ASAS']
    data['school_info'] = {
        'school_name': safe_val(ws.cell(5, 4).value) or 'SMK TUANKU LAILATUL SHAHREEN',
        'school_code': safe_val(ws.cell(5, 10).value) or 'REA0084',
        'principal': safe_val(ws.cell(7, 4).value) or 'LILI MARIAM BINTI MOHAMMAD @ MOKHTAR',
        'school_grade': safe_val(ws.cell(7, 10).value) or 'B',
        'phone': safe_val(ws.cell(9, 4).value) or '',
        'location': safe_val(ws.cell(9, 10).value) or '',
        'exam_year': safe_val(ws.cell(11, 4).value) or 2026,
        'zone': safe_val(ws.cell(11, 10).value) or '',
        'form': safe_val(ws.cell(13, 4).value) or 'LIMA',
        'email': safe_val(ws.cell(13, 10).value) or '',
        'su_name': safe_val(ws.cell(16, 10).value) or '',
        'su_phone': safe_val(ws.cell(17, 10).value) or '',
        'su_email': safe_val(ws.cell(18, 10).value) or '',
        'state': 'JABATAN PENDIDIKAN NEGERI PERLIS'
    }
    
    # Exam configuration (calon daftar/ambil)
    data['exam_config'] = {
        'tov': {'daftar': safe_val(ws.cell(16, 4).value), 'ambil': safe_val(ws.cell(16, 5).value)},
        'u1': {'daftar': safe_val(ws.cell(17, 4).value), 'ambil': safe_val(ws.cell(17, 5).value)},
        'ppt': {'daftar': safe_val(ws.cell(18, 4).value), 'ambil': safe_val(ws.cell(18, 5).value)},
        'spmc': {'daftar': safe_val(ws.cell(19, 4).value), 'ambil': safe_val(ws.cell(19, 5).value)},
        'kpi_2025': {'daftar': safe_val(ws.cell(21, 4).value), 'ambil': safe_val(ws.cell(21, 5).value)},
        'kpi_2024': {'daftar': safe_val(ws.cell(22, 4).value), 'ambil': safe_val(ws.cell(22, 5).value)},
        'kpi_2023': {'daftar': safe_val(ws.cell(23, 4).value), 'ambil': safe_val(ws.cell(23, 5).value)},
    }
    print(f"    School: {data['school_info']['school_name']}")
    print(f"    Code: {data['school_info']['school_code']}")
    
    # --- 2b. HC KESELURUHAN (Overall Performance) ---
    print("\n  [2b] HC KESELURUHAN (Overall Performance)...")
    ws = wb['HC KESELURUHAN']
    exams_hc = ['TOV', 'OTI1', 'U1', 'OTI2', 'PPT', 'OTI3', 'SPMC', 'ETR']
    rows_hc = [12, 13, 14, 15, 16, 17, 18, 19]
    for exam, row in zip(exams_hc, rows_hc):
        data['hc_keseluruhan'][exam] = {
            'calon_daftar': safe_val(ws.cell(row, 3).value),
            'calon_ambil': safe_val(ws.cell(row, 4).value),
            'cemerlang_a_bil': safe_val(ws.cell(row, 5).value),
            'cemerlang_a_pct': safe_val(ws.cell(row, 6).value),
            'layak_sijil_bil': safe_val(ws.cell(row, 7).value),
            'layak_sijil_pct': safe_val(ws.cell(row, 8).value),
            'gps': safe_val(ws.cell(row, 9).value),
        }
    # KPI years
    for yr_row in [20, 21, 22]:
        yr = safe_val(ws.cell(yr_row, 2).value)
        if yr:
            data['hc_keseluruhan'][f'KPI_{yr}'] = {
                'calon_daftar': safe_val(ws.cell(yr_row, 3).value),
                'calon_ambil': safe_val(ws.cell(yr_row, 4).value),
                'cemerlang_a_bil': safe_val(ws.cell(yr_row, 5).value),
                'cemerlang_a_pct': safe_val(ws.cell(yr_row, 6).value),
                'layak_sijil_bil': safe_val(ws.cell(yr_row, 7).value),
                'layak_sijil_pct': safe_val(ws.cell(yr_row, 8).value),
                'gps': safe_val(ws.cell(yr_row, 9).value),
            }
    print("    Done.")
    
    # --- 2c. HEADCOUNT (Overall Subject Headcount) ---
    print("\n  [2c] HEADCOUNT (Overall Subject Headcount)...")
    ws = wb['HEADCOUNT']
    grades = ['A+', 'A', 'A-', 'SEMUA_A', 'B+', 'B', 'C+', 'C', 'D', 'E', 'G', 'TH']
    grade_cols = [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
    
    exam_rows_hc = {
        'TOV': 12, 'OTI1': 14, 'U1': 16, 'OTI2': 18, 'PPT': 20,
        'OTI3': 22, 'SPMC': 24, 'ETR': 26, '2025': 28, '2024': 30, '2023': 32
    }
    
    for exam, row in exam_rows_hc.items():
        entry = {
            'calon_daftar': safe_val(ws.cell(row, 2).value),
            'calon_ambil': safe_val(ws.cell(row, 3).value),
            'lulus_bil': safe_val(ws.cell(row, 16).value),
            'gp': safe_val(ws.cell(row, 17).value),
            'grades': {},
            'grades_pct': {}
        }
        for g, c in zip(grades, grade_cols):
            entry['grades'][g] = safe_val(ws.cell(row, c).value)
            entry['grades_pct'][g] = safe_val(ws.cell(row + 1, c).value)
        lulus_pct = safe_val(ws.cell(row + 1, 16).value)
        entry['lulus_pct'] = lulus_pct
        data['headcount_overall'][exam] = entry
    print("    Done.")
    
    # --- 2d. ANALISIS - KPI ---
    print("\n  [2d] ANALISIS - KPI...")
    ws = wb['ANALISIS - KPI']
    kpi_subjects = []
    for row in range(12, 70):
        bil = safe_val(ws.cell(row, 1).value)
        bdg = safe_val(ws.cell(row, 2).value)
        subj = safe_val(ws.cell(row, 3).value)
        if subj and subj not in ['BAHASA', 'SAINS DAN MATEMATIK', 'KEMANUSIAAN', 'TEKNIK & VOKASIONAL', 'TEKNIK &VOKASIONAL']:
            entry = {
                'bil': bil, 'bidang': bdg, 'subject': subj,
                '2023_cln': safe_val(ws.cell(row, 4).value),
                '2023_a': safe_val(ws.cell(row, 5).value),
                '2023_pct': safe_val(ws.cell(row, 6).value),
                '2024_cln': safe_val(ws.cell(row, 7).value),
                '2024_a': safe_val(ws.cell(row, 8).value),
                '2024_pct': safe_val(ws.cell(row, 9).value),
                '2025_cln': safe_val(ws.cell(row, 10).value),
                '2025_a': safe_val(ws.cell(row, 11).value),
                '2025_pct': safe_val(ws.cell(row, 12).value),
                'etr_cln': safe_val(ws.cell(row, 13).value),
                'etr_a': safe_val(ws.cell(row, 14).value),
                'etr_pct': safe_val(ws.cell(row, 15).value),
                'lulus_2023': safe_val(ws.cell(row, 16).value),
                'lulus_2024': safe_val(ws.cell(row, 17).value),
                'lulus_2025': safe_val(ws.cell(row, 18).value),
                'lulus_etr_bil': safe_val(ws.cell(row, 19).value),
            }
            kpi_subjects.append(entry)
        elif bil in ['BAHASA', 'SAINS DAN MATEMATIK', 'KEMANUSIAAN', 'TEKNIK & VOKASIONAL', 'TEKNIK &VOKASIONAL'] or subj in ['BAHASA', 'SAINS DAN MATEMATIK', 'KEMANUSIAAN', 'TEKNIK & VOKASIONAL', 'TEKNIK &VOKASIONAL']:
            cat_name = bil if bil in ['BAHASA', 'SAINS DAN MATEMATIK', 'KEMANUSIAAN', 'TEKNIK & VOKASIONAL', 'TEKNIK &VOKASIONAL'] else subj
            entry = {
                'bil': 'SUBTOTAL', 'bidang': cat_name, 'subject': cat_name,
                '2023_cln': safe_val(ws.cell(row, 4).value),
                '2023_a': safe_val(ws.cell(row, 5).value),
                '2023_pct': safe_val(ws.cell(row, 6).value),
                '2024_cln': safe_val(ws.cell(row, 7).value),
                '2024_a': safe_val(ws.cell(row, 8).value),
                '2024_pct': safe_val(ws.cell(row, 9).value),
                '2025_cln': safe_val(ws.cell(row, 10).value),
                '2025_a': safe_val(ws.cell(row, 11).value),
                '2025_pct': safe_val(ws.cell(row, 12).value),
                'etr_cln': safe_val(ws.cell(row, 13).value),
                'etr_a': safe_val(ws.cell(row, 14).value),
                'etr_pct': safe_val(ws.cell(row, 15).value),
                'lulus_2023': safe_val(ws.cell(row, 16).value),
                'lulus_2024': safe_val(ws.cell(row, 17).value),
                'lulus_2025': safe_val(ws.cell(row, 18).value),
                'lulus_etr_bil': safe_val(ws.cell(row, 19).value),
            }
            kpi_subjects.append(entry)
    data['analysis_kpi'] = kpi_subjects
    print(f"    Extracted {len(kpi_subjects)} subject entries")
    
    # --- 2e. ANALISIS - HC ---
    print("\n  [2e] ANALISIS - HC...")
    ws = wb['ANALISIS - HC']
    hc_subjects = []
    for row in range(12, 70):
        bil = safe_val(ws.cell(row, 1).value)
        bdg = safe_val(ws.cell(row, 2).value)
        subj = safe_val(ws.cell(row, 3).value)
        if not subj and not bdg:
            continue
        entry = {
            'bil': bil, 'bidang': bdg, 'subject': subj or bil,
            'a_tov_bil': safe_val(ws.cell(row, 4).value),
            'a_tov_pct': safe_val(ws.cell(row, 5).value),
            'a_u1_bil': safe_val(ws.cell(row, 6).value),
            'a_u1_pct': safe_val(ws.cell(row, 7).value),
            'a_ppt_bil': safe_val(ws.cell(row, 8).value),
            'a_ppt_pct': safe_val(ws.cell(row, 9).value),
            'a_spmc_bil': safe_val(ws.cell(row, 10).value),
            'a_spmc_pct': safe_val(ws.cell(row, 11).value),
            'a_etr_bil': safe_val(ws.cell(row, 12).value),
            'a_etr_pct': safe_val(ws.cell(row, 13).value),
            'lulus_tov': safe_val(ws.cell(row, 14).value),
            'lulus_u1': safe_val(ws.cell(row, 15).value),
            'lulus_ppt': safe_val(ws.cell(row, 16).value),
            'lulus_spmc': safe_val(ws.cell(row, 17).value),
            'lulus_etr': safe_val(ws.cell(row, 18).value),
            'gpmp': safe_val(ws.cell(row, 19).value),
        }
        hc_subjects.append(entry)
    data['analysis_hc'] = hc_subjects
    print(f"    Extracted {len(hc_subjects)} subject entries")
    
    # --- 2f. A. CEMERLANG A ---
    print("\n  [2f] A. CEMERLANG A...")
    ws = wb['A. CEMERLANG A']
    cemerlang = {}
    exam_rows_ca = {'TOV': 12, 'OTI1': 13, 'U1': 14, 'OTI2': 15, 'PPT': 16, 'OTI3': 17, 'SPMC': 18, 'ETR': 19}
    for exam, row in exam_rows_ca.items():
        cemerlang[exam] = {
            'jumlah_calon': safe_val(ws.cell(row, 3).value),
            'semua_a_plus_bil': safe_val(ws.cell(row, 4).value),
            'semua_a_plus_pct': safe_val(ws.cell(row, 5).value),
            'a_plus_a_aminus_bil': safe_val(ws.cell(row, 6).value),
            'a_plus_a_aminus_pct': safe_val(ws.cell(row, 7).value),
            'jumlah_cemerlang_bil': safe_val(ws.cell(row, 8).value),
            'jumlah_cemerlang_pct': safe_val(ws.cell(row, 9).value),
            'near_miss_bplus_bil': safe_val(ws.cell(row, 10).value),
            'near_miss_bplus_pct': safe_val(ws.cell(row, 11).value),
            'near_miss_b_bil': safe_val(ws.cell(row, 12).value),
            'near_miss_b_pct': safe_val(ws.cell(row, 13).value),
            'jumlah_near_miss_bil': safe_val(ws.cell(row, 14).value),
            'jumlah_near_miss_pct': safe_val(ws.cell(row, 15).value),
        }
    # KPI years
    for yr_row in [20, 21, 22]:
        yr = safe_val(ws.cell(yr_row, 2).value)
        if yr:
            cemerlang[f'KPI_{yr}'] = {
                'jumlah_calon': safe_val(ws.cell(yr_row, 3).value),
                'semua_a_plus_pct': safe_val(ws.cell(yr_row, 5).value),
                'a_plus_a_aminus_bil': safe_val(ws.cell(yr_row, 6).value),
                'a_plus_a_aminus_pct': safe_val(ws.cell(yr_row, 7).value),
                'jumlah_cemerlang_bil': safe_val(ws.cell(yr_row, 8).value),
                'jumlah_cemerlang_pct': safe_val(ws.cell(yr_row, 9).value),
            }
    data['cemerlang_a'] = cemerlang
    print("    Done.")
    
    # --- 2g. A. LAYAK SIJIL ---
    print("\n  [2g] A. LAYAK SIJIL...")
    ws = wb['A. LAYAK SIJIL']
    layak = {}
    exam_rows_ls = {'TOV': 12, 'OTI1': 13, 'U1': 14, 'OTI2': 15, 'PPT': 16, 'OTI3': 17, 'SPMC': 18, 'ETR': 19}
    for exam, row in exam_rows_ls.items():
        layak[exam] = {
            'jumlah_calon': safe_val(ws.cell(row, 3).value),
            'lelaki': safe_val(ws.cell(row, 4).value),
            'perempuan': safe_val(ws.cell(row, 5).value),
            'layak_bil': safe_val(ws.cell(row, 6).value),
            'layak_pct': safe_val(ws.cell(row, 7).value),
            'layak_l': safe_val(ws.cell(row, 8).value),
            'layak_l_pct': safe_val(ws.cell(row, 9).value),
            'layak_p': safe_val(ws.cell(row, 10).value),
            'layak_p_pct': safe_val(ws.cell(row, 11).value),
            'tidak_layak_bil': safe_val(ws.cell(row, 12).value),
            'tidak_layak_pct': safe_val(ws.cell(row, 13).value),
            'tidak_layak_l': safe_val(ws.cell(row, 14).value),
            'tidak_layak_l_pct': safe_val(ws.cell(row, 15).value),
            'tidak_layak_p': safe_val(ws.cell(row, 16).value),
            'tidak_layak_p_pct': safe_val(ws.cell(row, 17).value),
            'gagal_bm_sej_bil': safe_val(ws.cell(row, 18).value),
            'gagal_bm_sej_pct': safe_val(ws.cell(row, 19).value),
        }
    for yr_row in [20, 21, 22]:
        yr = safe_val(ws.cell(yr_row, 2).value)
        if yr:
            layak[f'KPI_{yr}'] = {
                'jumlah_calon': safe_val(ws.cell(yr_row, 3).value),
                'layak_bil': safe_val(ws.cell(yr_row, 6).value),
                'layak_pct': safe_val(ws.cell(yr_row, 7).value),
                'tidak_layak_bil': safe_val(ws.cell(yr_row, 12).value),
                'tidak_layak_pct': safe_val(ws.cell(yr_row, 13).value),
            }
    data['layak_sijil'] = layak
    print("    Done.")
    
    # --- 2h. SUBJECT HEADCOUNT SHEETS (B1-B8, S1-S8, K1-K14, V1-V24) ---
    print("\n  [2h] Subject Headcount Sheets...")
    
    subject_sheets = {
        'bahasa': [f'B{i}' for i in range(1, 9)],
        'sains': [f'S{i}' for i in range(1, 9)],
        'kemanusiaan': [f'K{i}' for i in range(1, 15)],
        'vokasional': [f'V{i}' for i in range(1, 25)],
    }
    
    all_subjects = []
    
    for category, sheets in subject_sheets.items():
        data['subject_headcount'][category] = []
        for sn in sheets:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            subj_name = safe_val(ws.cell(8, 2).value) or ''
            if not subj_name or subj_name in ['G', 'GGG', 'H']:
                continue
            
            subject_data = {
                'sheet': sn,
                'category': category,
                'name': subj_name,
                'school': safe_val(ws.cell(6, 2).value),
                'form': safe_val(ws.cell(8, 16).value),
                'exams': {}
            }
            
            # Extract exam data for each exam period
            exam_rows_subj = {
                'TOV': 12, 'OTI1': 14, 'U1': 16, 'OTI2': 18, 'PPT': 20,
                'OTI3': 22, 'SPMC': 24, 'ETR': 26, '2025': 28, '2024': 30, '2023': 32
            }
            
            for exam, row in exam_rows_subj.items():
                exam_data = {
                    'calon_daftar': safe_val(ws.cell(row, 2).value),
                    'calon_ambil': safe_val(ws.cell(row, 3).value),
                    'semua_a': safe_val(ws.cell(row, 7).value),
                    'lulus_bil': safe_val(ws.cell(row, 16).value),
                    'gpmp': safe_val(ws.cell(row, 17).value),
                    'grades': {},
                    'grades_pct': {}
                }
                grade_names = ['A+', 'A', 'A-', 'B+', 'B', 'C+', 'C', 'D', 'E', 'G', 'TH']
                grade_cols_subj = [4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15]
                for g, c in zip(grade_names, grade_cols_subj):
                    exam_data['grades'][g] = safe_val(ws.cell(row, c).value)
                    exam_data['grades_pct'][g] = safe_val(ws.cell(row + 1, c).value)
                lulus_pct = safe_val(ws.cell(row + 1, 16).value)
                exam_data['lulus_pct'] = lulus_pct
                subject_data['exams'][exam] = exam_data
            
            data['subject_headcount'][category].append(subject_data)
            all_subjects.append({'sheet': sn, 'category': category, 'name': subj_name})
            print(f"    {sn}: {subj_name} ({category})")
    
    data['all_subjects_list'] = all_subjects
    
    # --- 2i. HC-TOV, HC-U1, HC-PPT, HC-SPMC, HC-ETR ---
    print("\n  [2i] HC by Exam Period sheets...")
    for hc_sheet in ['HC-TOV', 'HC-U1', 'HC-PPT', 'HC-SPMC', 'HC-ETR']:
        if hc_sheet not in wb.sheetnames:
            continue
        ws = wb[hc_sheet]
        exam_name = hc_sheet.replace('HC-', '')
        subjects_in_exam = []
        
        # Read all subject rows (start at row 12, every 2 rows)
        row = 12
        while row < 80:
            code = safe_val(ws.cell(row, 1).value)
            name = safe_val(ws.cell(row, 2).value)
            if not name:
                row += 2
                continue
            
            entry = {
                'code': code, 'name': name,
                'calon_daftar': safe_val(ws.cell(row, 3).value),
                'calon_ambil': safe_val(ws.cell(row, 4).value),
                'semua_a': safe_val(ws.cell(row, 8).value),
                'grades': {},
                'grades_pct': {}
            }
            g_names = ['A+', 'A', 'A-', 'B+', 'B', 'C+', 'C', 'D', 'E']
            g_cols = [5, 6, 7, 9, 10, 11, 12, 13, 14]
            for g, c in zip(g_names, g_cols):
                entry['grades'][g] = safe_val(ws.cell(row, c).value)
                entry['grades_pct'][g] = safe_val(ws.cell(row + 1, c).value)
            
            subjects_in_exam.append(entry)
            row += 2
        
        data['hc_by_exam'][exam_name] = subjects_in_exam
        print(f"    {hc_sheet}: {len(subjects_in_exam)} entries")
    
    # --- 2j. Individual Student Sheets (I1-I50) ---
    print("\n  [2j] Individual Student Sheets (I1-I50)...")
    subject_template = []
    for i in range(1, 51):
        sn = f'I{i}'
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        
        student_sheet = {
            'sheet': sn,
            'name': safe_val(ws.cell(7, 2).value) or '',
            'class': safe_val(ws.cell(7, 21).value) or '',
            'subjects': []
        }
        
        # Extract subjects (rows 12-23)
        for row in range(12, 24):
            subj_name = safe_val(ws.cell(row, 2).value)
            if not subj_name:
                continue
            
            subj = {
                'name': subj_name,
                'tov': {'marks': safe_val(ws.cell(row, 4).value), 'mata': safe_val(ws.cell(row, 5).value), 'gred': safe_val(ws.cell(row, 6).value)},
                'oti1': {'marks': safe_val(ws.cell(row, 7).value), 'mata': safe_val(ws.cell(row, 8).value), 'gred': safe_val(ws.cell(row, 9).value)},
                'u1': {'marks': safe_val(ws.cell(row, 10).value), 'mata': safe_val(ws.cell(row, 11).value), 'gred': safe_val(ws.cell(row, 12).value)},
                'oti2': {'marks': safe_val(ws.cell(row, 13).value), 'mata': safe_val(ws.cell(row, 14).value), 'gred': safe_val(ws.cell(row, 15).value)},
                'ppt': {'marks': safe_val(ws.cell(row, 16).value), 'mata': safe_val(ws.cell(row, 17).value), 'gred': safe_val(ws.cell(row, 18).value)},
                'oti3': {'marks': safe_val(ws.cell(row, 19).value), 'mata': safe_val(ws.cell(row, 20).value), 'gred': safe_val(ws.cell(row, 21).value)},
                'spmc': {'marks': safe_val(ws.cell(row, 22).value), 'mata': safe_val(ws.cell(row, 23).value), 'gred': safe_val(ws.cell(row, 24).value)},
                'etr': {'marks': safe_val(ws.cell(row, 25).value), 'mata': safe_val(ws.cell(row, 26).value), 'gred': safe_val(ws.cell(row, 27).value)},
            }
            student_sheet['subjects'].append(subj)
        
        # Summary rows
        student_sheet['jumlah_gred'] = {
            'tov': safe_val(ws.cell(24, 4).value), 'oti1': safe_val(ws.cell(24, 7).value),
            'u1': safe_val(ws.cell(24, 10).value), 'oti2': safe_val(ws.cell(24, 13).value),
            'ppt': safe_val(ws.cell(24, 16).value), 'oti3': safe_val(ws.cell(24, 19).value),
            'spmc': safe_val(ws.cell(24, 22).value), 'etr': safe_val(ws.cell(24, 25).value),
        }
        student_sheet['jumlah_mata'] = {
            'tov': safe_val(ws.cell(25, 4).value), 'oti1': safe_val(ws.cell(25, 7).value),
            'u1': safe_val(ws.cell(25, 10).value), 'oti2': safe_val(ws.cell(25, 13).value),
            'ppt': safe_val(ws.cell(25, 16).value), 'oti3': safe_val(ws.cell(25, 19).value),
            'spmc': safe_val(ws.cell(25, 22).value), 'etr': safe_val(ws.cell(25, 25).value),
        }
        student_sheet['jumlah_mp'] = {
            'tov': safe_val(ws.cell(26, 4).value), 'oti1': safe_val(ws.cell(26, 7).value),
            'u1': safe_val(ws.cell(26, 10).value), 'oti2': safe_val(ws.cell(26, 13).value),
            'ppt': safe_val(ws.cell(26, 16).value), 'oti3': safe_val(ws.cell(26, 19).value),
            'spmc': safe_val(ws.cell(26, 22).value), 'etr': safe_val(ws.cell(26, 25).value),
        }
        student_sheet['gred_purata'] = {
            'tov': safe_val(ws.cell(27, 4).value), 'oti1': safe_val(ws.cell(27, 7).value),
            'u1': safe_val(ws.cell(27, 10).value), 'oti2': safe_val(ws.cell(27, 13).value),
            'ppt': safe_val(ws.cell(27, 16).value), 'oti3': safe_val(ws.cell(27, 19).value),
            'spmc': safe_val(ws.cell(27, 22).value), 'etr': safe_val(ws.cell(27, 25).value),
        }
        
        # Get subject template from first sheet
        if i == 1:
            subject_template = [s['name'] for s in student_sheet['subjects']]
        
        data['individual_sheets'][sn] = student_sheet
    
    print(f"    Extracted {len(data['individual_sheets'])} individual sheets")
    print(f"    Subject template: {subject_template}")
    
    wb.close()
    
    # ================================================================
    # SAVE TO JSON
    # ================================================================
    output_file = 'spm_data.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    # ================================================================
    # VERIFICATION SUMMARY
    # ================================================================
    print("\n" + "="*70)
    print("EXTRACTION COMPLETE - VERIFICATION SUMMARY")
    print("="*70)
    print(f"\n  School: {data['school_info']['school_name']}")
    print(f"  Code: {data['school_info']['school_code']}")
    print(f"  Year: {data['school_info']['exam_year']}")
    print(f"  Principal: {data['school_info']['principal']}")
    print(f"\n  STUDENTS BY CLASS:")
    total = 0
    for cls, studs in data['students_by_class'].items():
        print(f"    {cls}: {len(studs)} students")
        total += len(studs)
    print(f"    TOTAL: {total} students")
    
    print(f"\n  SUBJECTS BY CATEGORY:")
    for cat, subjs in data['subject_headcount'].items():
        names = [s['name'] for s in subjs]
        print(f"    {cat.upper()} ({len(subjs)}): {', '.join(names)}")
    
    print(f"\n  ALL SUBJECTS LIST: {len(data['all_subjects_list'])}")
    print(f"  HC KESELURUHAN entries: {len(data['hc_keseluruhan'])}")
    print(f"  HEADCOUNT OVERALL entries: {len(data['headcount_overall'])}")
    print(f"  ANALYSIS KPI entries: {len(data['analysis_kpi'])}")
    print(f"  ANALYSIS HC entries: {len(data['analysis_hc'])}")
    print(f"  CEMERLANG A entries: {len(data['cemerlang_a'])}")
    print(f"  LAYAK SIJIL entries: {len(data['layak_sijil'])}")
    print(f"  HC BY EXAM entries: {len(data['hc_by_exam'])}")
    print(f"  INDIVIDUAL SHEETS: {len(data['individual_sheets'])}")
    
    print(f"\n  Data saved to: {output_file}")
    print("="*70)

if __name__ == "__main__":
    extract_all()
