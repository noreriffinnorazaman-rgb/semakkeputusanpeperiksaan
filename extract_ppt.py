"""
Extract Pentaksiran Pertengahan Tahun (PPT) marks from all 5 Tingkatan Excel files
into spm_data.json under marks_data[class][student]['PPT'].
Also stores ppt_data mirroring uat_data structure for headcount/analysis.
"""
import openpyxl
import json
import os

def safe_val(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    return s if s else None

def extract_ppt():
    data_file = 'spm_data.json'
    if os.path.exists(data_file):
        with open(data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    else:
        data = {}

    # Grade table for lower forms (Form 1-3)
    GRADE_TABLE_LOWER = [
        (80, 'A'), (70, 'B'), (60, 'C'), (50, 'D'), (40, 'E'), (0, 'F')
    ]
    # Grade table for upper forms (Form 4-5, SPM style)
    GRADE_TABLE_UPPER = [
        (90, 'A+', 1.0), (80, 'A', 2.0), (70, 'A-', 3.0),
        (65, 'B+', 4.0), (60, 'B', 5.0), (55, 'C+', 6.0),
        (50, 'C', 7.0), (45, 'D', 8.0), (40, 'E', 9.0), (0, 'G', 0.0)
    ]

    TAKSIRAN = {
        'A+': 'CEMERLANG', 'A': 'CEMERLANG', 'A-': 'KEPUJIAN TINGGI',
        'B+': 'KEPUJIAN ATAS', 'B': 'KEPUJIAN', 'C+': 'MEMUASKAN TINGGI',
        'C': 'BAIK', 'D': 'MEMUASKAN', 'E': 'MENCAPAI TAHAP MINIMUM',
        'F': 'BELUM MENCAPAI TAHAP MINIMUM', 'G': 'GAGAL', 'TH': 'TIDAK HADIR'
    }

    ppt_data = {}       # class_name -> list of student records (mirrors uat_data)
    ppt_guru_kelas = {} # class_name -> guru name

    for tingkatan in [1, 2, 3, 4, 5]:
        fn = f'PENTAKSIRAN PERTENGAHAN TAHUN TINGKATAN {tingkatan} 2026.xlsx'
        if not os.path.exists(fn):
            print(f'  SKIP: {fn} not found')
            continue

        print(f'\n{"="*60}')
        print(f'  EXTRACTING PPT TINGKATAN {tingkatan}: {fn}')
        print(f'{"="*60}')

        wb = openpyxl.load_workbook(fn, data_only=True)

        # --- Extract subject code->name mapping from SETUP ---
        ws_setup = wb['SETUP']
        subj_map = {}
        for r in range(2, 30):
            code = safe_val(ws_setup.cell(r, 1).value)
            name = safe_val(ws_setup.cell(r, 2).value)
            if code and name:
                name = name.replace('\n', ' ').strip()
                subj_map[code] = name
        print(f'  Subjects: {len(subj_map)} -> {list(subj_map.values())}')

        # --- Extract guru kelas from SETUP ---
        for r in range(2, 8):
            cls = safe_val(ws_setup.cell(r, 4).value)
            guru = safe_val(ws_setup.cell(r, 5).value)
            if cls and guru:
                ppt_guru_kelas[cls] = guru

        # --- Find class sheets ---
        class_sheets = [s for s in wb.sheetnames
                        if any(k in s for k in ['PROGRESIF', 'BESTARI', 'KREATIF', 'DINAMIK'])]
        print(f'  Class sheets: {class_sheets}')

        for sheet_name in class_sheets:
            ws = wb[sheet_name]

            # Build subject columns from row 2 headers
            headers = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(2, c).value
                if v:
                    headers[c] = str(v).strip()

            # Identify subject mark columns
            skip_names = {'NO', 'NO IC', 'NAMA PELAJAR', 'KELAS', 'JUMLAH MARKAH',
                          'JUM MARKAH', 'PURATA %', 'GP', 'KEPUTUSAN', 'KDK',
                          'ANALISIS', 'BIL', 'SORT', 'NO SORT', 'ULASAN',
                          'ULASAN GURU KELAS', 'KEHADIRAN'}
            grade_suffixes = set()
            for c, h in headers.items():
                if h.endswith(' G') or h in ['A+', 'A', 'A-', 'B+', 'B', 'C+', 'C', 'D', 'E', 'F', 'G', 'TH', 'TH+G']:
                    grade_suffixes.add(c)

            subject_cols = []  # list of (col, code, full_name)
            for c, h in sorted(headers.items()):
                if c in grade_suffixes:
                    continue
                if h in skip_names:
                    continue
                # Check if col+1 is a grade column
                grade_col = headers.get(c + 1, '')
                if grade_col.endswith(' G') or grade_col.endswith('G'):
                    full_name = subj_map.get(h, h)
                    subject_cols.append((c, h, full_name))

            # Identify summary columns
            sum_col_map = {}
            for c, h in headers.items():
                if h in ('JUMLAH MARKAH', 'JUM MARKAH'):
                    sum_col_map['jumlah_markah'] = c
                elif h == 'PURATA %':
                    sum_col_map['purata'] = c
                elif h == 'GP':
                    sum_col_map['gp'] = c
                elif h == 'KEPUTUSAN':
                    sum_col_map['keputusan'] = c
                elif h == 'KDK':
                    sum_col_map['kdk'] = c
                elif h == 'ULASAN GURU KELAS':
                    sum_col_map['ulasan'] = c
                elif h == 'KEHADIRAN':
                    sum_col_map['kehadiran'] = c
                elif h == 'ANALISIS':
                    sum_col_map['analisis'] = c

            class_subject_names = [sc[2] for sc in subject_cols]

            print(f'\n  {sheet_name}: {len(subject_cols)} subjects')
            print(f'    Subjects: {[sc[1] for sc in subject_cols]}')

            # --- Extract student data ---
            students = []
            for row in range(3, ws.max_row + 1):
                no = safe_val(ws.cell(row, 1).value)
                name = safe_val(ws.cell(row, 3).value)
                if no is None or name is None:
                    continue
                try:
                    int(float(str(no)))
                except (ValueError, TypeError):
                    continue

                ic_no = safe_val(ws.cell(row, 2).value)
                kelas = safe_val(ws.cell(row, 4).value) or sheet_name

                student = {
                    'no': int(float(str(no))),
                    'ic': str(ic_no) if ic_no else '',
                    'name': name,
                    'class': kelas,
                    'form': f'T{tingkatan}',
                    'tingkatan': tingkatan,
                    'subjects': {},
                    'subject_order': class_subject_names,
                }

                total_marks = 0
                total_gp = 0
                subj_count = 0

                for col, code, full_name in subject_cols:
                    mark = safe_val(ws.cell(row, col).value)
                    grade = safe_val(ws.cell(row, col + 1).value)

                    if mark == 'TH' or grade == 'TH':
                        student['subjects'][full_name] = {
                            'mark': 'TH', 'grade': 'TH',
                            'taksiran': 'TIDAK HADIR', 'gp': None
                        }
                        continue

                    if mark is not None and mark != '':
                        try:
                            m = float(mark)
                            if grade is None or grade == '':
                                # Compute grade
                                if tingkatan >= 4:
                                    for thresh, g, gp in GRADE_TABLE_UPPER:
                                        if m >= thresh:
                                            grade = g
                                            break
                                else:
                                    for thresh, g in GRADE_TABLE_LOWER:
                                        if m >= thresh:
                                            grade = g
                                            break

                            # Compute GP
                            gp = None
                            if tingkatan >= 4:
                                for thresh, g, p in GRADE_TABLE_UPPER:
                                    if m >= thresh:
                                        gp = p
                                        break
                            else:
                                gp_map = {'A': 1.0, 'B': 2.0, 'C': 3.0, 'D': 4.0, 'E': 5.0, 'F': 6.0}
                                gp = gp_map.get(str(grade), None)

                            taksiran = TAKSIRAN.get(str(grade), '')
                            student['subjects'][full_name] = {
                                'mark': m, 'grade': str(grade), 'taksiran': taksiran, 'gp': gp
                            }
                            total_marks += m
                            if gp is not None:
                                total_gp += gp
                                subj_count += 1
                        except (ValueError, TypeError):
                            student['subjects'][full_name] = {
                                'mark': str(mark), 'grade': str(grade) if grade else '',
                                'taksiran': '', 'gp': None
                            }
                    else:
                        student['subjects'][full_name] = {
                            'mark': '', 'grade': '', 'taksiran': '', 'gp': None
                        }

                # Summary
                student['jumlah_markah'] = safe_val(ws.cell(row, sum_col_map.get('jumlah_markah', 0)).value) if 'jumlah_markah' in sum_col_map else total_marks
                student['purata'] = safe_val(ws.cell(row, sum_col_map.get('purata', 0)).value) if 'purata' in sum_col_map else (round(total_marks / subj_count, 2) if subj_count > 0 else 0)
                excel_gp = safe_val(ws.cell(row, sum_col_map.get('gp', 0)).value) if 'gp' in sum_col_map else None
                student['gp'] = round(float(excel_gp), 4) if excel_gp and excel_gp not in ('TH',) else (round(total_gp / subj_count, 4) if subj_count > 0 else 99.0)
                student['keputusan'] = safe_val(ws.cell(row, sum_col_map.get('keputusan', 0)).value) if 'keputusan' in sum_col_map else ''
                student['analisis'] = safe_val(ws.cell(row, sum_col_map.get('analisis', 0)).value) if 'analisis' in sum_col_map else ''
                student['ulasan'] = safe_val(ws.cell(row, sum_col_map.get('ulasan', 0)).value) if 'ulasan' in sum_col_map else ''
                student['kehadiran'] = safe_val(ws.cell(row, sum_col_map.get('kehadiran', 0)).value) if 'kehadiran' in sum_col_map else ''
                student['bil_subjects'] = subj_count

                students.append(student)

            ppt_data[sheet_name] = students
            print(f'    Extracted {len(students)} students')

        wb.close()

    # ================================================================
    # COMPUTE RANKINGS (Kedudukan Dalam Kelas & Tingkatan)
    # ================================================================
    print(f'\n{"="*60}')
    print('  COMPUTING PPT RANKINGS...')
    print(f'{"="*60}')

    tingkatan_groups = {}
    for cls_name, students in ppt_data.items():
        for s in students:
            t = s['tingkatan']
            tingkatan_groups.setdefault(t, []).append(s)

    # Rank within each class
    for cls_name, students in ppt_data.items():
        valid = [s for s in students if s['gp'] < 90]
        valid.sort(key=lambda x: (x['gp'], -x.get('jumlah_markah', 0) if isinstance(x.get('jumlah_markah'), (int, float)) else 0))
        total_in_class = len(students)
        rank = 1
        for i, s in enumerate(valid):
            if i > 0 and valid[i]['gp'] != valid[i-1]['gp']:
                rank = i + 1
            s['kdk'] = f'{rank}/{total_in_class}'
            s['kdk_rank'] = rank
        for s in students:
            if 'kdk' not in s:
                s['kdk'] = f'-/{total_in_class}'
                s['kdk_rank'] = 9999

    # Rank within each tingkatan
    for t, all_students in tingkatan_groups.items():
        valid = [s for s in all_students if s['gp'] < 90]
        valid.sort(key=lambda x: (x['gp'], -x.get('jumlah_markah', 0) if isinstance(x.get('jumlah_markah'), (int, float)) else 0))
        total_in_form = len(all_students)
        rank = 1
        for i, s in enumerate(valid):
            if i > 0 and valid[i]['gp'] != valid[i-1]['gp']:
                rank = i + 1
            s['kdt'] = f'{rank}/{total_in_form}'
            s['kdt_rank'] = rank
        for s in all_students:
            if 'kdt' not in s:
                s['kdt'] = f'-/{total_in_form}'
                s['kdt_rank'] = 9999

    # ================================================================
    # SAVE TO JSON — update marks_data with PPT marks
    # ================================================================
    print(f'\n{"="*60}')
    print('  SAVING PPT DATA...')
    print(f'{"="*60}')

    # Store ppt_data as a parallel structure to uat_data
    data['ppt_data'] = ppt_data

    # Update marks_data for PPT exam
    if 'marks_data' not in data:
        data['marks_data'] = {}

    # First, match PPT student names to UAT student names for compatibility
    uat_data = data.get('uat_data', {})

    for cls_name, students in ppt_data.items():
        if cls_name not in data['marks_data']:
            data['marks_data'][cls_name] = {}

        # Build UAT name lookup for this class
        uat_names = set()
        if cls_name in uat_data:
            for s in uat_data[cls_name]:
                uat_names.add(s.get('name', ''))

        for s in students:
            sname = s['name']

            if sname not in data['marks_data'][cls_name]:
                data['marks_data'][cls_name][sname] = {}
            data['marks_data'][cls_name][sname]['PPT'] = {}

            for subj_name, subj_data in s['subjects'].items():
                mark = subj_data.get('mark', '')
                data['marks_data'][cls_name][sname]['PPT'][subj_name] = mark

            # If student is NOT in uat_data for this class, we need to add them
            if sname not in uat_names:
                print(f'    NEW student in PPT (not in UAT): {cls_name} -> {sname}')
                # Add to uat_data so they appear in the system
                if cls_name not in uat_data:
                    uat_data[cls_name] = []
                uat_data[cls_name].append({
                    'no': s['no'],
                    'ic': s['ic'],
                    'name': sname,
                    'class': s['class'],
                    'form': s['form'],
                    'tingkatan': s['tingkatan'],
                    'subjects': {subj: {'mark': '', 'grade': '', 'taksiran': '', 'gp': None}
                                 for subj in s['subject_order']},
                    'subject_order': s['subject_order'],
                    'jumlah_markah': 0,
                    'purata': 0,
                    'gp': 99.0,
                    'keputusan': '',
                    'analisis': '',
                    'ulasan': '',
                    'kehadiran': '',
                    'bil_subjects': 0,
                    'kdk': '-/-',
                    'kdk_rank': 9999,
                    'kdt': '-/-',
                    'kdt_rank': 9999,
                })

    # Update uat_data in case new students were added
    data['uat_data'] = uat_data

    # Also update guru_kelas if any new ones
    existing_guru = data.get('guru_kelas', {})
    existing_guru.update(ppt_guru_kelas)
    data['guru_kelas'] = existing_guru

    with open(data_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    # ================================================================
    # VERIFICATION
    # ================================================================
    print(f'\n{"="*60}')
    print('  PPT EXTRACTION COMPLETE - VERIFICATION')
    print(f'{"="*60}')
    total = 0
    for cls_name in sorted(ppt_data.keys()):
        students = ppt_data[cls_name]
        total += len(students)
        top = min(students, key=lambda x: x.get('gp', 99)) if students else None
        if top:
            print(f'  {cls_name}: {len(students)} students | Top: {top["name"][:30]} (GP={top["gp"]:.2f}, KDK={top["kdk"]}, KDT={top["kdt"]})')
        else:
            print(f'  {cls_name}: {len(students)} students')
    print(f'\n  TOTAL: {total} students across {len(ppt_data)} classes')

    # Verify marks_data PPT entries
    ppt_mark_count = 0
    for cls_name, students_dict in data.get('marks_data', {}).items():
        for sname, exams in students_dict.items():
            if 'PPT' in exams:
                ppt_mark_count += 1
    print(f'  PPT marks_data entries: {ppt_mark_count}')
    print(f'  Data saved to: {data_file}')

if __name__ == '__main__':
    extract_ppt()
