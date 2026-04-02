import openpyxl

wb2 = openpyxl.load_workbook("SENARAI NAMA MURID2026.xlsx", data_only=True)

# Get remaining T5 data (5 KREATIF continued + any more classes)
ws = wb2['T5']
print("T5 REMAINING (rows 115-200):")
for row in range(115, min(ws.max_row + 1, 250)):
    vals = []
    for col in range(1, 7):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# Get remaining T4 data
ws = wb2['T4']
print("\nT4 REMAINING (rows 115-200):")
for row in range(115, min(ws.max_row + 1, 250)):
    vals = []
    for col in range(1, 7):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# Get remaining T3 data
ws = wb2['T3']
print("\nT3 REMAINING (rows 118-175):")
for row in range(118, min(ws.max_row + 1, 200)):
    vals = []
    for col in range(1, 7):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# Get remaining T2 data
ws = wb2['T2']
print("\nT2 REMAINING (rows 105-200):")
for row in range(105, min(ws.max_row + 1, 250)):
    vals = []
    for col in range(1, 7):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

# Get remaining T1 data
ws = wb2['T1']
print("\nT1 REMAINING (rows 118-169):")
for row in range(118, min(ws.max_row + 1, 200)):
    vals = []
    for col in range(1, 7):
        v = ws.cell(row, col).value
        if v is not None:
            vals.append(f"C{col}={v}")
    if vals:
        print(f"Row {row}: {vals}")

wb2.close()
